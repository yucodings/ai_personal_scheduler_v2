from io import BytesIO
from openpyxl import load_workbook

def parse_xlsx(content: bytes) -> str:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True); output = []
    try:
        for worksheet in workbook.worksheets:
            output.append(f"[Sheet: {worksheet.title}]")
            for row in worksheet.iter_rows():
                values = [str(cell.value) if cell.value is not None else "" for cell in row]
                if any(values): output.append(" | ".join(values).rstrip(" |"))
    finally: workbook.close()
    return "\n".join(output)

